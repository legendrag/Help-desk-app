from django.core.exceptions import ValidationError, PermissionDenied
from django.http import HttpResponse
from datetime import datetime, timedelta, time
from collections import defaultdict
import os
from django.utils import timezone as tz

from django.views.generic import ListView, DetailView, CreateView, UpdateView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages as django_messages
from django.urls import reverse_lazy
from django.db.models import Count, Q, F, ExpressionWrapper, fields, Avg
# TruncDate, TruncMonth, TruncWeek, TruncYear removed because DB-side timezone conversion crashes SQLite with USE_TZ=True
from .models import Ticket, TicketMessage, TicketStatusHistory
from core.models import Branch, Category, Department
from accounts.models import User
from .forms import TicketCreateForm, TicketUpdateForm
from .services import merge_tickets
from notifications.services import notify_ticket_picked, notify_ticket_update
# ... (existing code)

def _default_month_range():
    """Return (start_date, end_date) for the current local calendar month."""
    today = tz.localdate()
    start = today.replace(day=1)
    end = (start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    return start, end


def _aware_day_bounds(start_date, end_date):
    """
    Convert local calendar dates to aware datetimes for filtering.

    Avoid created_at__date lookups: with USE_TZ=True they rely on DB CONVERT_TZ,
    which returns NULL on MySQL when timezone tables are missing and empties results.
    """
    start_dt = None
    end_dt = None
    if start_date:
        start_dt = tz.make_aware(datetime.combine(start_date, time.min))
    if end_date:
        # exclusive upper bound: start of the next local day
        end_dt = tz.make_aware(datetime.combine(end_date + timedelta(days=1), time.min))
    return start_dt, end_dt


def _iter_bucket_keys(start_date, end_date, view):
    """Yield continuous bucket keys between start_date and end_date for date_view."""
    if not start_date or not end_date or start_date > end_date:
        return

    if view == "year":
        for year in range(start_date.year, end_date.year + 1):
            yield str(year)
    elif view == "month":
        current = start_date.replace(day=1)
        end_month = end_date.replace(day=1)
        while current <= end_month:
            yield current.strftime("%Y-%m")
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)
    elif view == "week":
        current = start_date - timedelta(days=start_date.weekday())
        end_monday = end_date - timedelta(days=end_date.weekday())
        while current <= end_monday:
            yield current.strftime("%Y-%m-%d")
            current += timedelta(days=7)
    else:
        current = start_date
        while current <= end_date:
            yield current.strftime("%Y-%m-%d")
            current += timedelta(days=1)


def _format_bucket_label(key, view):
    if view == "year":
        return key
    if view == "month":
        return datetime.strptime(key, "%Y-%m").strftime("%b %Y")
    if view == "week":
        return f"Week of {datetime.strptime(key, '%Y-%m-%d').strftime('%m/%d/%Y')}"
    return datetime.strptime(key, "%Y-%m-%d").strftime("%m/%d/%Y")


def _status_breakdown_rows(queryset, group_field):
    """
    Build rows like {label, count, statuses[{value, label, count, percent}]}
    grouped by group_field (e.g. branch__name / department__name).
    """
    status_labels = dict(Ticket.Status.choices)
    status_order = [value for value, _ in Ticket.Status.choices]

    grouped = list(
        queryset.values(group_field, "status")
        .annotate(count=Count("id"))
        .order_by(group_field, "status")
    )

    by_label = defaultdict(lambda: defaultdict(int))
    for item in grouped:
        label = item[group_field] or "—"
        by_label[label][item["status"]] += item["count"]

    rows = []
    for label, status_counts in by_label.items():
        total = sum(status_counts.values())
        statuses = []
        for value in status_order:
            count = status_counts.get(value, 0)
            if count <= 0:
                continue
            percent = int(round((count / total) * 100)) if total else 0
            statuses.append(
                {
                    "value": value,
                    "label": status_labels.get(value, value),
                    "count": count,
                    "percent": percent,
                }
            )
        # Fix rounding so percents sum to ~100 for flex basis
        if statuses and total:
            drift = 100 - sum(s["percent"] for s in statuses)
            if drift:
                statuses[0]["percent"] = max(0, statuses[0]["percent"] + drift)

        rows.append({"label": label, "count": total, "statuses": statuses})

    rows.sort(key=lambda row: (-row["count"], row["label"] or ""))
    return rows


class SettingsView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = "core/settings.html"

    def test_func(self):
        user = self.request.user
        if user.is_superuser:
            return True
        return user.role and user.role.can_access_settings

class DashboardView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Ticket
    template_name = "tickets/dashboard.html"
    context_object_name = "tickets"

    def test_func(self):
        user = self.request.user
        if user.is_superuser:
            return True
        return user.role and user.role.can_access_dashboard

    def get_template_names(self):
        if self.request.user.user_type == "branch":
            return ["tickets/branch_dashboard.html"]
        return [self.template_name]

    def get_queryset(self):
        user = self.request.user
        queryset = Ticket.objects.select_related(
            "branch", "department", "category", "assigned_to", "created_by"
        ).all()

        if not user.is_superuser:
            if user.user_type == "branch" and user.branch_id:
                queryset = queryset.filter(branch_id=user.branch_id)
            elif user.user_type == "support" and user.department_id:
                queryset = queryset.filter(department_id=user.department_id)
            else:
                queryset = queryset.none()
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        def parse_date(value):
            if not value:
                return None
            try:
                return datetime.strptime(value, "%Y-%m-%d").date()
            except ValueError:
                return None

        base_queryset = self.get_queryset()

        filters = {
            "start_date": self.request.GET.get("start_date", ""),
            "end_date": self.request.GET.get("end_date", ""),
            "department": self.request.GET.get("department", "all"),
            "branch": self.request.GET.get("branch", "all"),
            "assigned_to": self.request.GET.get("assigned_to", "all"),
            "date_view": self.request.GET.get("date_view", "day"),
        }

        if filters["date_view"] not in {"day", "week", "month", "year"}:
            filters["date_view"] = "day"

        filtered_queryset = base_queryset
        start_date = parse_date(filters["start_date"])
        end_date = parse_date(filters["end_date"])

        if not start_date and not end_date:
            start_date, end_date = _default_month_range()
            filters["start_date"] = start_date.isoformat()
            filters["end_date"] = end_date.isoformat()

        start_dt, end_dt = _aware_day_bounds(start_date, end_date)
        if start_dt:
            filtered_queryset = filtered_queryset.filter(created_at__gte=start_dt)
        if end_dt:
            filtered_queryset = filtered_queryset.filter(created_at__lt=end_dt)

        if filters["department"] and filters["department"] != "all":
            filtered_queryset = filtered_queryset.filter(department_id=filters["department"])
        if filters["branch"] and filters["branch"] != "all":
            filtered_queryset = filtered_queryset.filter(branch_id=filters["branch"])
        if filters["assigned_to"] and filters["assigned_to"] != "all":
            if filters["assigned_to"] == "unassigned":
                filtered_queryset = filtered_queryset.filter(assigned_to__isnull=True)
            else:
                filtered_queryset = filtered_queryset.filter(assigned_to_id=filters["assigned_to"])

        status_counts = {
            item["status"]: item["count"]
            for item in filtered_queryset.values("status").annotate(count=Count("id"))
        }

        status_summary = []
        for value, label in Ticket.Status.choices:
            status_summary.append(
                {
                    "value": value,
                    "label": label,
                    "count": status_counts.get(value, 0),
                }
            )

        def apply_percent(items):
            max_count = max((item["count"] for item in items), default=0)
            for item in items:
                item["percent"] = int((item["count"] / max_count) * 100) if max_count else 0
            return items

        department_items = _status_breakdown_rows(filtered_queryset, "department__name")
        branch_items = _status_breakdown_rows(filtered_queryset, "branch__name")

        category_items = list(
            filtered_queryset.values("category__name", "category__department__name")
            .annotate(count=Count("id"))
            .order_by("-count", "category__name")
        )
        category_items = [
            {
                "label": item["category__name"], 
                "department": item["category__department__name"],
                "count": item["count"]
            } for item in category_items
        ]

        def bucket_key(dt, view):
            local_dt = tz.localtime(dt)
            if view == "year":
                return local_dt.strftime("%Y")
            elif view == "month":
                return local_dt.strftime("%Y-%m")
            elif view == "week":
                monday = local_dt.date() - timedelta(days=local_dt.weekday())
                return monday.strftime("%Y-%m-%d")
            else:
                return local_dt.strftime("%Y-%m-%d")

        raw_timestamps = filtered_queryset.values_list("created_at", flat=True)
        bucket_counts = defaultdict(int)
        for ts in raw_timestamps:
            if ts:
                bucket_counts[bucket_key(ts, filters["date_view"])] += 1

        view = filters["date_view"]
        if start_date and end_date:
            ordered_keys = list(_iter_bucket_keys(start_date, end_date, view))
        else:
            ordered_keys = sorted(bucket_counts.keys())

        formatted_dates = [
            {"label": _format_bucket_label(key, view), "count": bucket_counts.get(key, 0)}
            for key in ordered_keys
        ]

        total_tickets = filtered_queryset.count()
        total_users = User.objects.filter(
            Q(created_tickets__in=filtered_queryset) | Q(assigned_tickets__in=filtered_queryset)
        ).distinct().count()
        branches_count = filtered_queryset.values("branch_id").distinct().count()
        departments_count = filtered_queryset.values("department_id").distinct().count()

        if user.is_superuser:
            departments = Department.objects.all()
            branches = Branch.objects.all()
        elif user.user_type == "branch" and user.branch_id:
            departments = Department.objects.filter(tickets__branch_id=user.branch_id).distinct()
            branches = Branch.objects.filter(id=user.branch_id)
        elif user.user_type == "support" and user.department_id:
            departments = Department.objects.filter(id=user.department_id)
            branches = Branch.objects.filter(tickets__department_id=user.department_id).distinct()
        else:
            departments = Department.objects.none()
            branches = Branch.objects.none()

        assignee_ids = base_queryset.exclude(assigned_to__isnull=True).values_list("assigned_to_id", flat=True)

        def build_query(date_view):
            params = self.request.GET.copy()
            if date_view:
                params["date_view"] = date_view
            else:
                params.pop("date_view", None)
            return params.urlencode()

        drill_down_view = None
        drill_up_view = None
        if filters["date_view"] == "year":
            drill_down_view = "month"
        elif filters["date_view"] == "month":
            drill_down_view = "week"
            drill_up_view = "year"
        elif filters["date_view"] == "week":
            drill_down_view = "day"
            drill_up_view = "month"
        else:
            drill_up_view = "week"

        # Performance Leaderboard
        agent_performance = []
        if user.is_superuser or (user.role and getattr(user.role, 'can_view_leaderboard', False)):
            resolved_tickets = filtered_queryset.filter(status=Ticket.Status.CLOSED, assigned_to__isnull=False)

            performance_qs = resolved_tickets.values(
                'assigned_to__username', 'assigned_to__first_name', 'assigned_to__last_name'
            ).annotate(
                tickets_resolved=Count('id'),
                avg_response_time=Avg(
                    ExpressionWrapper(F('picked_at') - F('created_at'), output_field=fields.DurationField())
                ),
                avg_resolution_time=Avg(
                    ExpressionWrapper(F('closed_at') - F('created_at'), output_field=fields.DurationField())
                )
            ).order_by('-tickets_resolved')[:10]

            # Working time = sum of (closed_at - picked_at - waiting) across resolved tickets.
            # Computed in Python so pending seconds subtract cleanly on SQLite and MySQL.
            working_totals = defaultdict(float)
            for username, picked_at, closed_at, pending in resolved_tickets.exclude(
                picked_at__isnull=True
            ).exclude(closed_at__isnull=True).values_list(
                "assigned_to__username",
                "picked_at",
                "closed_at",
                "total_pending_duration_seconds",
            ):
                working_seconds = (closed_at - picked_at).total_seconds() - (pending or 0)
                if working_seconds < 0:
                    working_seconds = 0
                working_totals[username] += working_seconds

            def format_duration(duration):
                if not duration:
                    return "--"
                total_seconds = int(duration.total_seconds()) if hasattr(duration, "total_seconds") else int(duration)
                if total_seconds < 0:
                    total_seconds = 0
                hours, remainder = divmod(total_seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                if hours > 24:
                    days, hours = divmod(hours, 24)
                    return f"{days}d {hours}h {minutes}m"
                return f"{hours}h {minutes}m"

            def total_working_duration(username):
                total_seconds = working_totals.get(username)
                if not total_seconds:
                    return None
                return timedelta(seconds=total_seconds)

            TARGET_VOLUME = 30 # Loosened volume baseline
            MAX_HOURS = 40 # Loosened worst acceptable avg resolution time

            for agent in performance_qs:
                name = f"{agent['assigned_to__first_name']} {agent['assigned_to__last_name']}".strip() or agent['assigned_to__username']
                username = agent['assigned_to__username']

                resolved = agent['tickets_resolved']
                res_time = agent['avg_resolution_time']
                work_time = total_working_duration(username)

                vol_score = min(100, (resolved / TARGET_VOLUME) * 100) if TARGET_VOLUME else 0

                if res_time:
                    res_hours = res_time.total_seconds() / 3600
                    speed_score = max(0, 100 - ((res_hours / MAX_HOURS) * 100))
                else:
                    speed_score = 100

                final_score = round((speed_score * 0.8) + (vol_score * 0.2))

                if final_score >= 85:
                    grade, color = 'A', 'success'
                elif final_score >= 70:
                    grade, color = 'B', 'primary'
                elif final_score >= 55:
                    grade, color = 'C', 'warning'
                elif final_score >= 40:
                    grade, color = 'D', 'warning'
                else:
                    grade, color = 'F', 'danger'

                agent_performance.append({
                    "name": name,
                    "tickets_resolved": resolved,
                    "avg_response_time": format_duration(agent['avg_response_time']),
                    "avg_resolution_time": format_duration(res_time),
                    "total_working_time": format_duration(work_time),
                    "score": final_score,
                    "grade": grade,
                    "grade_color": color
                })

            # Sort by highest score first
            agent_performance.sort(key=lambda x: x['score'], reverse=True)

        # Recent Activity Feed
        recent_activity = base_queryset.order_by('-updated_at')[:5]

        context.update(
            {
                "filter_values": filters,
                "departments": departments,
                "branches": branches,
                "assignees": User.objects.filter(id__in=assignee_ids).order_by("username"),
                "total_tickets": total_tickets,
                "total_users": total_users,
                "branches_count": branches_count,
                "departments_count": departments_count,
                "status_summary": apply_percent(status_summary),
                "tickets_by_department": department_items,
                "tickets_by_category": apply_percent(category_items),
                "tickets_by_branch": branch_items,
                "status_legend": [
                    {"value": value, "label": label} for value, label in Ticket.Status.choices
                ],
                "tickets_by_date": apply_percent(formatted_dates),
                "date_view": filters["date_view"],
                "drill_down_query": build_query(drill_down_view) if drill_down_view else "",
                "drill_up_query": build_query(drill_up_view) if drill_up_view else "",
                "agent_performance": agent_performance,
                "recent_activity": recent_activity,
            }
        )

        return context

import openpyxl
from django.utils import timezone
from openpyxl.styles import Font, Alignment

class ExportDashboardExcelView(DashboardView):
    """Reuses DashboardView logic to build an Excel workbook with multiple sheets."""
    def get(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        
        def parse_date(value):
            if not value: return None
            try: return datetime.strptime(value, "%Y-%m-%d").date()
            except ValueError: return None
            
        filters = {
            "start_date": self.request.GET.get("start_date", ""),
            "end_date": self.request.GET.get("end_date", ""),
            "department": self.request.GET.get("department", "all"),
            "branch": self.request.GET.get("branch", "all"),
            "assigned_to": self.request.GET.get("assigned_to", "all"),
            "date_view": self.request.GET.get("date_view", "month")
        }
        
        qs = self.object_list
        start_date = parse_date(filters["start_date"])
        end_date = parse_date(filters["end_date"])

        if not start_date and not end_date:
            start_date, end_date = _default_month_range()
            filters["start_date"] = start_date.isoformat()
            filters["end_date"] = end_date.isoformat()

        start_dt, end_dt = _aware_day_bounds(start_date, end_date)
        if start_dt:
            qs = qs.filter(created_at__gte=start_dt)
        if end_dt:
            qs = qs.filter(created_at__lt=end_dt)
        if filters["department"] and filters["department"] != "all":
            qs = qs.filter(department_id=filters["department"])
        if filters["branch"] and filters["branch"] != "all":
            qs = qs.filter(branch_id=filters["branch"])
        if filters["assigned_to"] and filters["assigned_to"] != "all":
            if filters["assigned_to"] == "unassigned":
                qs = qs.filter(assigned_to__isnull=True)
            else:
                qs = qs.filter(assigned_to_id=filters["assigned_to"])

        context = self.get_context_data(object_list=self.object_list)
        
        wb = openpyxl.Workbook()
        
        # Helper for headers
        def make_header(ws, row, headers):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        # Sheet 1: Tickets
        ws_tickets = wb.active
        ws_tickets.title = "Tickets"
        make_header(ws_tickets, 1, ['Ticket Number', 'Title', 'Branch', 'Department', 'Category', 'Status', 'Priority', 'Created By', 'Assigned To', 'Created At', 'Closed At'])
        
        for row, ticket in enumerate(qs, 2):
            ws_tickets.cell(row=row, column=1, value=ticket.ticket_number)
            ws_tickets.cell(row=row, column=2, value=ticket.title)
            ws_tickets.cell(row=row, column=3, value=ticket.branch.name if ticket.branch else '')
            ws_tickets.cell(row=row, column=4, value=ticket.department.name if ticket.department else '')
            ws_tickets.cell(row=row, column=5, value=ticket.category.name if ticket.category else '')
            ws_tickets.cell(row=row, column=6, value=ticket.get_status_display())
            ws_tickets.cell(row=row, column=7, value=ticket.get_priority_display())
            ws_tickets.cell(row=row, column=8, value=ticket.created_by.username if ticket.created_by else '')
            ws_tickets.cell(row=row, column=9, value=ticket.assigned_to.username if ticket.assigned_to else 'Unassigned')
            ws_tickets.cell(row=row, column=10, value=ticket.created_at.strftime("%Y-%m-%d %H:%M") if ticket.created_at else '')
            ws_tickets.cell(row=row, column=11, value=ticket.closed_at.strftime("%Y-%m-%d %H:%M") if ticket.closed_at else '')

        # Sheet 2: Agent Performance
        ws_perf = wb.create_sheet(title="Agent Performance")
        make_header(ws_perf, 1, [
            'Agent',
            'Resolved',
            'Avg Response Time',
            'Avg Resolution Time',
            'Total Working Time',
            'Score',
            'Grade',
        ])
        for row, agent in enumerate(context.get('agent_performance', []), 2):
            ws_perf.cell(row=row, column=1, value=agent['name'])
            ws_perf.cell(row=row, column=2, value=agent['tickets_resolved'])
            ws_perf.cell(row=row, column=3, value=agent['avg_response_time'])
            ws_perf.cell(row=row, column=4, value=agent['avg_resolution_time'])
            ws_perf.cell(row=row, column=5, value=agent['total_working_time'])
            ws_perf.cell(row=row, column=6, value=agent['score'])
            ws_perf.cell(row=row, column=7, value=agent['grade'])

        # Sheet 3: Status Summary
        ws_status = wb.create_sheet(title="Status Summary")
        make_header(ws_status, 1, ['Status', 'Count', 'Percentage'])
        for row, stat in enumerate(context.get('status_summary', []), 2):
            ws_status.cell(row=row, column=1, value=stat['label'])
            ws_status.cell(row=row, column=2, value=stat['count'])
            ws_status.cell(row=row, column=3, value=f"{stat.get('percent', 0)}%")

        status_headers = [label for _, label in Ticket.Status.choices]

        def write_status_breakdown_sheet(ws, title_col, rows):
            make_header(ws, 1, [title_col, *status_headers, 'Total'])
            for row_num, item in enumerate(rows, 2):
                counts_by_status = {
                    s["value"]: s["count"] for s in item.get("statuses", [])
                }
                ws.cell(row=row_num, column=1, value=item["label"])
                for col_num, (value, _) in enumerate(Ticket.Status.choices, 2):
                    ws.cell(row=row_num, column=col_num, value=counts_by_status.get(value, 0))
                ws.cell(row=row_num, column=2 + len(status_headers), value=item["count"])

        # Sheet 4: Departments
        ws_dept = wb.create_sheet(title="Departments")
        write_status_breakdown_sheet(
            ws_dept, "Department", context.get("tickets_by_department", [])
        )

        # Sheet 5: Categories
        ws_cat = wb.create_sheet(title="Categories")
        make_header(ws_cat, 1, ['Category', 'Department', 'Count', 'Percentage'])
        for row, cat in enumerate(context.get('tickets_by_category', []), 2):
            ws_cat.cell(row=row, column=1, value=cat['label'])
            ws_cat.cell(row=row, column=2, value=cat.get('department', ''))
            ws_cat.cell(row=row, column=3, value=cat['count'])
            ws_cat.cell(row=row, column=4, value=f"{cat.get('percent', 0)}%")

        # Sheet 6: Branches
        ws_branch = wb.create_sheet(title="Branches")
        write_status_breakdown_sheet(
            ws_branch, "Branch", context.get("tickets_by_branch", [])
        )

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="dashboard_export_{timezone.now().strftime("%Y%m%d%H%M")}.xlsx"'
        wb.save(response)
        
        response.set_cookie('fileDownload', 'true', path='/')
        return response

class TicketCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Ticket
    form_class = TicketCreateForm
    template_name = "tickets/create.html"
    success_url = reverse_lazy("tickets_list")

    def get_success_url(self):
        if self.request.user.user_type == "branch":
            from django.urls import reverse
            return reverse("ticket_detail", kwargs={"ticket_id": self.object.id}) + "#chat-box"
        return super().get_success_url()
    def test_func(self):
        user = self.request.user
        if user.is_superuser:
            return True
        return user.role and user.role.can_create_ticket

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        django_messages.success(self.request, "Ticket created successfully.")

        TicketStatusHistory.objects.create(ticket=self.object, status=self.object.status, changed_by=self.request.user)

        
        if self.request.headers.get('HX-Request'):
            from django.http import HttpResponse
            from django.urls import reverse
            resp = HttpResponse(status=204)
            if self.request.user.user_type == "branch":
                resp['HX-Redirect'] = reverse("ticket_detail", kwargs={"ticket_id": self.object.id}) + "#chat-box"
            else:
                resp['HX-Trigger'] = 'closeModal,refreshTickets'
            return resp
            
        return response

    def get_template_names(self):
        if self.request.headers.get('HX-Request'):
            return ["tickets/create_partial.html"]
        return [self.template_name]


class TicketUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Ticket
    form_class = TicketUpdateForm
    template_name = "tickets/edit.html"
    success_url = reverse_lazy("tickets_list")
    pk_url_kwarg = "ticket_id"

    def test_func(self):
        user = self.request.user
        if user.is_superuser:
            return True
            
        ticket_id = self.kwargs.get(self.pk_url_kwarg)
        if ticket_id:
            ticket = Ticket.objects.filter(id=ticket_id).first()
            if ticket and ticket.assigned_to_id == user.id:
                return True
        
        return user.role and user.role.can_update_ticket

    def get_queryset(self):
        return Ticket.objects.select_related(
            "branch", "department", "category", "created_by", "assigned_to"
        ).all()

    def get_object(self, queryset=None):
        ticket = super().get_object(queryset)
        user = self.request.user

        if user.is_superuser:
            return ticket

        if user.user_type == "branch":
            if ticket.branch_id != user.branch_id:
                raise PermissionDenied("You do not have permission to edit this ticket.")
        elif user.user_type == "support":
            if ticket.department_id != user.department_id:
                raise PermissionDenied("You do not have permission to edit this ticket.")
        else:
            raise PermissionDenied("You do not have permission to edit this ticket.")

        return ticket

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        django_messages.success(self.request, "Ticket updated successfully.")

        if "priority" in form.changed_data:
            old_priority = form.initial.get("priority")
            new_priority = form.cleaned_data.get("priority")
            TicketStatusHistory.objects.create(
                ticket=self.object,
                event_type=TicketStatusHistory.EventType.PRIORITY_CHANGED,
                detail=f"{old_priority} → {new_priority}" if old_priority else f"set to {new_priority}",
                changed_by=self.request.user,
            )
        if "assigned_to" in form.changed_data:
            new_assignee = form.cleaned_data.get("assigned_to")
            TicketStatusHistory.objects.create(
                ticket=self.object,
                event_type=TicketStatusHistory.EventType.ASSIGNED,
                detail=new_assignee.username if new_assignee else "Unassigned",
                changed_by=self.request.user,
            )

        notify_ticket_update(self.object, self.request.user)

        if self.request.headers.get('HX-Request'):
            from django.http import HttpResponse
            resp = HttpResponse(status=204)
            resp['HX-Trigger'] = 'closeModal,refreshTickets'
            return resp

        return response

    def get_template_names(self):
        if self.request.headers.get('HX-Request'):
            return ["tickets/edit_partial.html"]
        return [self.template_name]

class TicketListView(LoginRequiredMixin, ListView):
    # ... (existing code)
    model = Ticket
    template_name = "tickets/list.html"
    context_object_name = "tickets"
    paginate_by = 10

    def get_template_names(self):
        if self.request.headers.get('HX-Request'):
            if self.request.GET.get('append') == 'true':
                return ["tickets/list_append.html"]
            return ["tickets/list_live_partial.html"]
        return [self.template_name]

    def get_queryset(self):
        user = self.request.user
        queryset = Ticket.objects.select_related(
            "branch", "department", "category", "created_by", "assigned_to"
        ).all()

        if user.is_superuser:
            base_queryset = queryset
        elif user.user_type == "branch":
            base_queryset = queryset.filter(branch_id=user.branch_id)
        elif user.user_type == "support":
            base_queryset = queryset.filter(department_id=user.department_id)
        else:
            base_queryset = queryset.none()

        self.base_queryset = base_queryset

        search_query = self.request.GET.get("q", "").strip()
        if search_query:
            base_queryset = base_queryset.filter(
                Q(ticket_number__icontains=search_query)
                | Q(title__icontains=search_query)
                | Q(description__icontains=search_query)
            )

        branch_filter = self.request.GET.get("branch")
        if branch_filter and branch_filter != "all":
            base_queryset = base_queryset.filter(branch_id=branch_filter)

        status_filter = self.request.GET.get("status")
        if status_filter and status_filter != "all":
            base_queryset = base_queryset.filter(status=status_filter)

        assignment_filter = self.request.GET.get("assignment")
        if assignment_filter and assignment_filter != "all":
            if assignment_filter == "unassigned":
                base_queryset = base_queryset.filter(assigned_to__isnull=True)
            else:
                base_queryset = base_queryset.filter(assigned_to_id=assignment_filter)

        return base_queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        context["filter_values"] = {
            "q": self.request.GET.get("q", ""),
            "branch": self.request.GET.get("branch", "all"),
            "status": self.request.GET.get("status", "all"),
            "assignment": self.request.GET.get("assignment", "all"),
        }

        if user.is_superuser:
            branches = Branch.objects.all()
        elif user.user_type == "branch" and user.branch_id:
            branches = Branch.objects.filter(id=user.branch_id)
        elif user.user_type == "support" and user.department_id:
            branches = Branch.objects.filter(tickets__department_id=user.department_id).distinct()
        else:
            branches = Branch.objects.none()

        base_queryset = getattr(self, "base_queryset", Ticket.objects.none())
        assignee_ids = base_queryset.exclude(assigned_to__isnull=True).values_list("assigned_to_id", flat=True).distinct()

        context["branches"] = branches
        context["assignees"] = User.objects.filter(id__in=assignee_ids).order_by("username")
        context["status_choices"] = Ticket.Status.choices

        params = self.request.GET.copy()
        params.pop("page", None)
        context["filters_query"] = params.urlencode()

        from news.models import Announcement
        from django.utils import timezone
        
        now = timezone.now()
        announcements = Announcement.objects.filter(is_active=True).filter(
            Q(expires_at__isnull=True) | Q(expires_at__gt=now)
        )
        
        if not user.is_superuser:
            if user.user_type == "branch":
                announcements = announcements.filter(
                    Q(target_branch__isnull=True) | Q(target_branch=user.branch)
                )
            elif user.user_type == "support":
                announcements = announcements.filter(
                    target_branch__isnull=True
                )
                
        context["active_announcements"] = announcements.order_by('-created_at')

        return context

class TicketDetailView(LoginRequiredMixin, DetailView):
    model = Ticket
    template_name = "tickets/detail.html"
    context_object_name = "ticket"
    pk_url_kwarg = "ticket_id"

    def get_queryset(self):
        return Ticket.objects.select_related(
            "branch", "department", "category", "created_by", "assigned_to"
        ).prefetch_related("messages", "messages__sender", "status_history", "status_history__changed_by")

    def get_object(self, queryset=None):
        ticket = super().get_object(queryset)
        user = self.request.user

        if user.is_superuser:
            return ticket

        has_kb_access = ticket.kb_articles.filter(is_published=True).exists()

        if user.user_type == "branch":
            if not (ticket.branch_id == user.branch_id or has_kb_access):
                raise PermissionDenied("You do not have permission to view this ticket.")
        elif user.user_type == "support":
            if not (ticket.department_id == user.department_id or has_kb_access):
                raise PermissionDenied("You do not have permission to view this ticket.")
        else:
            raise PermissionDenied("You do not have permission to view this ticket.")

        return ticket

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ticket = self.get_object()
        user = self.request.user

        # Calculate metrics
        context['response_time'] = int((ticket.picked_at - ticket.created_at).total_seconds()) if ticket.picked_at else None
        if ticket.closed_at and ticket.picked_at:
            duration = (ticket.closed_at - ticket.picked_at).total_seconds()
            context['resolution_time'] = int(duration - ticket.total_pending_duration_seconds)
        else:
            context['resolution_time'] = None
        context['time_to_close'] = int((ticket.closed_at - ticket.created_at).total_seconds()) if ticket.closed_at else None

        can_chat = False
        if user.is_superuser:
            can_chat = True
        elif user.user_type == "branch":
            can_chat = (ticket.branch_id == user.branch_id)
        elif user.user_type == "support":
            can_chat = (ticket.assigned_to_id == user.id)
        context['can_chat'] = can_chat

        supporters = User.objects.filter(
            status=User.Status.ACTIVE
        ).filter(
            Q(department_id=ticket.department_id, user_type=User.UserType.SUPPORT) |
            Q(is_superuser=True)
        ).exclude(id=ticket.assigned_to_id)

        context['supporters'] = supporters

        return context


class TicketDrawerPartialView(LoginRequiredMixin, DetailView):
    model = Ticket
    template_name = "tickets/partials/details_drawer.html"
    context_object_name = "ticket"
    pk_url_kwarg = "ticket_id"

    def get_queryset(self):
        return Ticket.objects.select_related(
            "branch", "department", "category", "assigned_to"
        ).prefetch_related(
            "merged_tickets", "kb_articles", "kb_articles__created_by",
            "status_history", "status_history__changed_by"
        )

    def get_object(self, queryset=None):
        ticket = super().get_object(queryset)
        user = self.request.user

        if user.is_superuser:
            return ticket

        has_kb_access = ticket.kb_articles.filter(is_published=True).exists()

        if user.user_type == "branch":
            if not (ticket.branch_id == user.branch_id or has_kb_access):
                raise PermissionDenied("You do not have permission to view this ticket.")
        elif user.user_type == "support":
            if not (ticket.department_id == user.department_id or has_kb_access):
                raise PermissionDenied("You do not have permission to view this ticket.")
        else:
            raise PermissionDenied("You do not have permission to view this ticket.")

        return ticket

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ticket = self.get_object()
        user = self.request.user

        # Calculate metrics (same as detail view)
        context['response_time'] = int((ticket.picked_at - ticket.created_at).total_seconds()) if ticket.picked_at else None
        if ticket.closed_at and ticket.picked_at:
            duration = (ticket.closed_at - ticket.picked_at).total_seconds()
            context['resolution_time'] = int(duration - ticket.total_pending_duration_seconds)
        else:
            context['resolution_time'] = None
        context['time_to_close'] = int((ticket.closed_at - ticket.created_at).total_seconds()) if ticket.closed_at else None

        # Add auto_open context to trigger automatic opening when fetched via HTMX
        context['auto_open'] = True

        return context

# PermissionDenied imported at the top

@login_required
def ticket_category_options(request):
    user = request.user
    if not user.is_superuser:
        if user.user_type == "support":
            if not (user.role and (user.role.can_create_ticket or user.role.can_update_ticket)):
                raise PermissionDenied("You do not have permission to access categories.")

    department_id = request.GET.get("department")
    if department_id:
        categories = Category.objects.filter(department_id=department_id).order_by("name")
    else:
        categories = Category.objects.none()

    return render(request, "tickets/category_options.html", {"categories": categories})


def post_message(request, ticket_id):
    ticket = get_object_or_404(Ticket, pk=ticket_id)
    
    # Permission: Superuser, Assignee, Branch (if theirs), or has role permission.
    if not request.user.is_superuser:
        if request.user.user_type == "branch":
            if ticket.branch_id != request.user.branch_id:
                raise PermissionDenied("You can only send messages on tickets for your branch.")
        elif request.user.user_type == "support":
            if ticket.assigned_to_id != request.user.id and not (request.user.role and request.user.role.can_send_message):
                raise PermissionDenied("You do not have permission to send messages.")

    if request.method == "POST":
        message_text = request.POST.get("message")
        attachment = request.FILES.get("attachment")
        reply_to_id = request.POST.get("reply_to")

        if not message_text and not attachment:
            django_messages.error(request, "Message or attachment is required.")
        else:
            try:
                reply_to = None
                if reply_to_id:
                    reply_to = TicketMessage.objects.filter(id=reply_to_id, ticket=ticket).first()
                msg = TicketMessage.objects.create(
                    ticket=ticket,
                    sender=request.user,
                    message=message_text,
                    attachment=attachment,
                    reply_to=reply_to
                )

            except ValidationError as e:
                for message in e.messages:
                    django_messages.error(request, message)
            except Exception as e:
                django_messages.error(request, f"Failed to send message: {str(e)}")

        # For HTMX requests, return 204 to avoid full page reload
        if request.headers.get('HX-Request'):
            from django.http import HttpResponse
            return HttpResponse(status=204)

    return redirect("ticket_detail", ticket_id=ticket_id)

def delete_message(request, message_id):
    message = get_object_or_404(TicketMessage, pk=message_id)
    if not (request.user.is_superuser or (request.user.role and request.user.role.can_delete_message and message.sender == request.user)):
        raise PermissionDenied("You do not have permission to delete this message.")
    ticket_id = message.ticket.id
    msg_id = message.id
    if request.method == "POST":
        message.delete()
        django_messages.success(request, "Message deleted.")

        # Broadcast delete event via WebSocket
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"ticket_{ticket_id}",
            {
                "type": "chat.event",
                "event": "message_deleted",
                "payload": {"id": msg_id},
            },
        )

        if request.headers.get('HX-Request'):
            from django.http import HttpResponse
            return HttpResponse(status=204)

    return redirect("ticket_detail", ticket_id=ticket_id)

def edit_message(request, message_id):
    message = get_object_or_404(TicketMessage, pk=message_id)
    if not (request.user.is_superuser or (request.user.role and request.user.role.can_edit_message and message.sender == request.user)):
        raise PermissionDenied("You do not have permission to edit this message.")
    ticket_id = message.ticket.id
    if request.method == "POST":
        new_text = request.POST.get("message")
        if new_text:
            message.message = new_text
            message.save()
            django_messages.success(request, "Message updated.")

            # Broadcast edit event via WebSocket
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"ticket_{ticket_id}",
                {
                    "type": "chat.event",
                    "event": "message_edited",
                    "payload": {
                        "id": message.id,
                        "message": message.message,
                        "updated_at": message.updated_at.isoformat(),
                    },
                },
            )

            if request.headers.get('HX-Request'):
                from django.http import HttpResponse
                return HttpResponse(status=204)
        else:
            django_messages.error(request, "Message text cannot be empty.")
    return redirect("ticket_detail", ticket_id=ticket_id)

def _broadcast_ticket_update(ticket, event_name, extra_payload=None):
    """Broadcast a ticket event to both the detail page and list page WebSocket groups."""
    from channels.layers import get_channel_layer
    from asgiref.sync import async_to_sync

    channel_layer = get_channel_layer()
    payload = {
        "id": ticket.id,
        "ticket_number": ticket.ticket_number,
        "status": ticket.status,
        "status_display": ticket.get_status_display(),
        "assigned_to": ticket.assigned_to.username if ticket.assigned_to else None,
        "assigned_to_id": ticket.assigned_to_id,
    }
    if extra_payload:
        payload.update(extra_payload)

    # 1) Broadcast to the ticket detail page (TicketChatConsumer group)
    async_to_sync(channel_layer.group_send)(
        f"ticket_{ticket.id}",
        {
            "type": "chat.event",
            "event": event_name,
            "payload": payload,
        },
    )

    # 2) Broadcast to ticket list groups (TicketListConsumer groups)
    list_groups = ["ticket_list"]  # superadmin group
    if ticket.branch_id:
        list_groups.append(f"ticket_list_branch_{ticket.branch_id}")
    if ticket.department_id:
        list_groups.append(f"ticket_list_department_{ticket.department_id}")

    for group_name in list_groups:
        async_to_sync(channel_layer.group_send)(
            group_name,
            {
                "type": "ticket.event",
                "event": event_name,
                "payload": payload,
            },
        )


def update_ticket_status(request, ticket_id):
    if request.method == "POST":
        ticket = get_object_or_404(Ticket, pk=ticket_id)
        
        if ticket.status != Ticket.Status.CLOSED:
            if not request.user.is_superuser and ticket.assigned_to_id != request.user.id:
                raise PermissionDenied("You can only update the status of tickets assigned to you.")
        else:
            if not (request.user.is_superuser or request.user.user_type == 'support' or (request.user.role and request.user.role.can_update_closed_ticket)):
                raise PermissionDenied("You do not have permission to update a closed ticket.")

        new_status = request.POST.get("status")
        if new_status in Ticket.Status.values:
            try:
                if ticket.status != new_status:
                    old_status = ticket.status
                    if ticket.status == Ticket.Status.CLOSED and new_status != Ticket.Status.CLOSED:
                        ticket.assigned_to = request.user
                        
                    ticket.status = new_status
                    
                    if new_status in [Ticket.Status.CLOSED, Ticket.Status.MERGED]:
                        ticket.pending_transfer_to = None
                        ticket.pending_transfer_by = None
                        
                    ticket.save()
                    
                    if old_status == Ticket.Status.CLOSED and new_status != Ticket.Status.CLOSED:
                        TicketStatusHistory.objects.create(
                            ticket=ticket,
                            status=new_status,
                            event_type=TicketStatusHistory.EventType.REOPENED,
                            changed_by=request.user,
                        )
                    else:
                        TicketStatusHistory.objects.create(
                            ticket=ticket,
                            status=new_status,
                            event_type=TicketStatusHistory.EventType.STATUS_CHANGE,
                            changed_by=request.user,
                        )
                    django_messages.success(request, f"Status updated to {ticket.get_status_display()}.")
                    notify_ticket_update(ticket, request.user, status_changed=True, new_status=new_status)

                    # Broadcast status change via WebSocket
                    _broadcast_ticket_update(ticket, "ticket_status_changed", {
                        "changed_by": request.user.username,
                        "new_status": new_status,
                        "new_status_display": ticket.get_status_display(),
                    })
                
                if request.headers.get('HX-Request'):
                    from django.http import HttpResponse
                    resp = HttpResponse(status=204)
                    resp['HX-Trigger'] = 'reloadPage'
                    return resp
            except ValidationError as e:
                for message in e.messages:
                    django_messages.error(request, message)
            except Exception as e:
                django_messages.error(request, f"Failed to update status: {str(e)}")
    return redirect("ticket_detail", ticket_id=ticket_id)

def pick_ticket(request, ticket_id):
    if not (request.user.is_superuser or (request.user.role and request.user.role.can_pick_ticket)):
        raise PermissionDenied("You do not have permission to pick tickets.")

    if request.method == "POST":
        ticket = get_object_or_404(Ticket, pk=ticket_id)
        if not ticket.assigned_to:
            if ticket.status == Ticket.Status.MERGED:
                django_messages.error(request, "Cannot pick a merged ticket.")
            else:
                try:
                    ticket.assigned_to = request.user
                    ticket.status = Ticket.Status.IN_PROGRESS
                    ticket.save()
                    TicketStatusHistory.objects.create(
                        ticket=ticket,
                        status=Ticket.Status.IN_PROGRESS,
                        changed_by=request.user,
                    )
                    django_messages.success(request, "Ticket assigned to you.")
                    notify_ticket_picked(ticket, request.user)

                    # Broadcast pick event via WebSocket
                    _broadcast_ticket_update(ticket, "ticket_picked", {
                        "picked_by": request.user.username,
                    })

                except ValidationError as e:
                    for message in e.messages:
                        django_messages.error(request, message)
                except Exception as e:
                    django_messages.error(request, f"Failed to pick ticket: {str(e)}")

    if request.headers.get('HX-Request'):
        from django.http import HttpResponse
        resp = HttpResponse(status=204)
        resp['HX-Trigger'] = 'reloadPage,refreshTickets'
        return resp

    return redirect("ticket_detail", ticket_id=ticket_id)

def transfer_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, pk=ticket_id)
    if not (request.user.is_superuser or ticket.assigned_to_id == request.user.id):
        raise PermissionDenied("You do not have permission to transfer this ticket.")

    if request.method == "POST":
        new_assignee_id = request.POST.get("new_assignee")
        if new_assignee_id:
            try:
                new_assignee = User.objects.get(id=new_assignee_id, status=User.Status.ACTIVE)
                
                # Only allow transfer to same-department supporters or superusers
                is_same_dept = (new_assignee.user_type == User.UserType.SUPPORT 
                                and new_assignee.department_id == ticket.department_id)
                if not (new_assignee.is_superuser or is_same_dept):
                    django_messages.error(request, "You can only transfer to agents in the same department or admins.")
                elif ticket.pending_transfer_to:
                    django_messages.error(request, "A transfer is already pending.")
                else:
                    ticket.pending_transfer_to = new_assignee
                    ticket.pending_transfer_by = request.user
                    ticket.save()
                    
                    TicketMessage.objects.create(
                        ticket=ticket,
                        sender=request.user,
                        message=f"Requested transfer to {new_assignee.username}",
                        is_system_message=True
                    )
                    
                    TicketStatusHistory.objects.create(
                        ticket=ticket,
                        event_type=TicketStatusHistory.EventType.TRANSFER_REQUESTED,
                        detail=f"→ {new_assignee.username}",
                        changed_by=request.user,
                    )
                    
                    from notifications.services import notify_transfer_requested
                    notify_transfer_requested(ticket, request.user, new_assignee)
                    
                    django_messages.success(request, f"Transfer request sent to {new_assignee.username}.")
                    
                    _broadcast_ticket_update(ticket, "ticket_transfer_update")
                    
            except User.DoesNotExist:
                django_messages.error(request, "Selected user is not a valid supporter.")
            except Exception as e:
                django_messages.error(request, f"Failed to request transfer: {str(e)}")

    if request.headers.get('HX-Request'):
        from django.http import HttpResponse
        resp = HttpResponse(status=204)
        resp['HX-Trigger'] = 'reloadPage,refreshTickets'
        return resp

    return redirect("ticket_detail", ticket_id=ticket_id)

@login_required
def accept_transfer(request, ticket_id):
    ticket = get_object_or_404(Ticket, pk=ticket_id)
    if request.user != ticket.pending_transfer_to:
        raise PermissionDenied("You do not have permission to accept this transfer.")
        
    if request.method == "POST":
        requester = ticket.pending_transfer_by
        ticket.assigned_to = request.user
        ticket.pending_transfer_to = None
        ticket.pending_transfer_by = None
        ticket.save()
        
        TicketMessage.objects.create(
            ticket=ticket,
            sender=request.user,
            message="Accepted ticket transfer",
            is_system_message=True
        )
        
        TicketStatusHistory.objects.create(
            ticket=ticket,
            event_type=TicketStatusHistory.EventType.TRANSFER_ACCEPTED,
            detail=f"from {requester.username}" if requester else "",
            changed_by=request.user,
        )
        
        from notifications.services import notify_transfer_accepted
        if requester:
            notify_transfer_accepted(ticket, request.user, requester)
            
        django_messages.success(request, "Ticket transfer accepted.")
        
        _broadcast_ticket_update(ticket, "ticket_picked", {
            "picked_by": request.user.username,
        })
        _broadcast_ticket_update(ticket, "ticket_transfer_update")
        
    if request.headers.get('HX-Request'):
        from django.http import HttpResponse
        resp = HttpResponse(status=204)
        resp['HX-Trigger'] = 'reloadPage,refreshTickets'
        return resp
    return redirect("ticket_detail", ticket_id=ticket_id)

@login_required
def deny_transfer(request, ticket_id):
    ticket = get_object_or_404(Ticket, pk=ticket_id)
    if request.user != ticket.pending_transfer_to:
        raise PermissionDenied("You do not have permission to deny this transfer.")
        
    if request.method == "POST":
        requester = ticket.pending_transfer_by
        ticket.pending_transfer_to = None
        ticket.pending_transfer_by = None
        ticket.save()
        
        TicketMessage.objects.create(
            ticket=ticket,
            sender=request.user,
            message="Denied ticket transfer",
            is_system_message=True
        )
        
        TicketStatusHistory.objects.create(
            ticket=ticket,
            event_type=TicketStatusHistory.EventType.TRANSFER_DENIED,
            detail=f"from {requester.username}" if requester else "",
            changed_by=request.user,
        )
        
        from notifications.services import notify_transfer_denied
        if requester:
            notify_transfer_denied(ticket, request.user, requester)
            
        django_messages.success(request, "Ticket transfer denied.")
        _broadcast_ticket_update(ticket, "ticket_transfer_update")
        
    if request.headers.get('HX-Request'):
        from django.http import HttpResponse
        resp = HttpResponse(status=204)
        resp['HX-Trigger'] = 'reloadPage,refreshTickets'
        return resp
    return redirect("ticket_detail", ticket_id=ticket_id)

@login_required
def cancel_transfer(request, ticket_id):
    ticket = get_object_or_404(Ticket, pk=ticket_id)
    if not (request.user.is_superuser or request.user == ticket.pending_transfer_by):
        raise PermissionDenied("You do not have permission to cancel this transfer.")
        
    if request.method == "POST":
        target = ticket.pending_transfer_to
        ticket.pending_transfer_to = None
        ticket.pending_transfer_by = None
        ticket.save()
        
        TicketMessage.objects.create(
            ticket=ticket,
            sender=request.user,
            message=f"Canceled ticket transfer to {target.username if target else 'unknown'}",
            is_system_message=True
        )
        
        TicketStatusHistory.objects.create(
            ticket=ticket,
            event_type=TicketStatusHistory.EventType.TRANSFER_CANCELLED,
            detail=f"to {target.username}" if target else "",
            changed_by=request.user,
        )
        
        django_messages.success(request, "Ticket transfer canceled.")
        _broadcast_ticket_update(ticket, "ticket_transfer_update")
        
    if request.headers.get('HX-Request'):
        from django.http import HttpResponse
        resp = HttpResponse(status=204)
        resp['HX-Trigger'] = 'reloadPage,refreshTickets'
        return resp
    return redirect("ticket_detail", ticket_id=ticket_id)

from django.db.models import Q
from django.contrib.auth.decorators import login_required

@login_required
def ticket_search_options(request):
    user = request.user
    query = request.GET.get("q", "") or request.GET.get("target_ticket_number", "")
    query = query.strip()
    exclude_id = request.GET.get("exclude")
    
    if len(query) < 1:
        return HttpResponse('<div id="search-results-container"></div>')
        
    tickets = Ticket.objects.filter(
        Q(ticket_number__icontains=query) | Q(title__icontains=query)
    ).exclude(merged_into__isnull=False)
    
    if not user.is_superuser:
        if user.user_type == "branch" and user.branch_id:
            tickets = tickets.filter(branch_id=user.branch_id)
        elif user.user_type == "support" and user.department_id:
            tickets = tickets.filter(department_id=user.department_id)
        else:
            tickets = tickets.none()
    
    if exclude_id and exclude_id.isdigit():
        tickets = tickets.exclude(id=exclude_id)
        
    tickets = tickets.select_related("department", "branch", "created_by")[:10]
    
    if not tickets.exists():
        return HttpResponse('<div id="search-results-container" class="merge-search-results"><div class="merge-search-empty">No matching tickets found</div></div>')

    options = ['<div id="search-results-container" class="merge-search-results">']
    for t in tickets:
        status_label = t.get_status_display()
        priority_label = t.get_priority_display().upper()
        options.append(
            f'<div class="merge-search-item" '
            f'hx-get="/tickets/merge-preview/{t.id}/" '
            f'hx-target="#selected-ticket-preview" '
            f'hx-swap="innerHTML" '
            f'onclick="selectTicketForMerge(\'{t.ticket_number}\')">'
            f'  <div class="merge-item-header">'
            f'    <span class="merge-item-number">{t.ticket_number}</span>'
            f'    <div class="merge-item-badges">'
            f'      <span class="badge badge-{t.status}">{status_label}</span>'
            f'      <span class="priority-badge priority-{t.priority}">{priority_label}</span>'
            f'    </div>'
            f'  </div>'
            f'  <div class="merge-item-title">{t.title}</div>'
            f'  <div class="merge-item-meta">'
            f'    <span>Dept: {t.department.name}</span> | '
            f'    <span>By: {t.created_by.username}</span>'
            f'  </div>'
            f'</div>'
        )
    options.append('</div>')
        
    return HttpResponse("\n".join(options))

@login_required
def ticket_merge_preview(request, ticket_id):
    user = request.user
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    if not user.is_superuser:
        if user.user_type == "branch" and ticket.branch_id != user.branch_id:
            raise PermissionDenied("You do not have permission to view this ticket.")
        elif user.user_type == "support" and ticket.department_id != user.department_id:
            raise PermissionDenied("You do not have permission to view this ticket.")
            
    msg_count = ticket.messages.count()
    return render(request, "tickets/merge_preview_partial.html", {
        "ticket": ticket,
        "msg_count": msg_count,
    })

def merge_ticket(request, ticket_id):
    if not (request.user.is_superuser or (request.user.role and request.user.role.can_update_status)):
        raise PermissionDenied("You do not have permission to merge tickets.")

    primary_ticket = get_object_or_404(Ticket, pk=ticket_id)

    if request.method == "POST":
        target_ticket_number = request.POST.get("target_ticket_number")
        if not target_ticket_number:
            django_messages.error(request, "Target ticket number is required.")
            return redirect("ticket_detail", ticket_id=ticket_id)

        try:
            secondary_ticket = Ticket.objects.get(ticket_number=target_ticket_number)
        except Ticket.DoesNotExist:
            django_messages.error(request, f"Ticket {target_ticket_number} not found.")
            return redirect("ticket_detail", ticket_id=ticket_id)

        try:
            merge_tickets(primary_ticket.id, [secondary_ticket.id], request.user)
            django_messages.success(request, f"Ticket {secondary_ticket.ticket_number} merged into this ticket successfully.")
            
            # Broadcast the secondary ticket status change as 'merged'
            _broadcast_ticket_update(secondary_ticket, "ticket_status_changed", {
                "changed_by": request.user.username,
                "new_status": "merged",
                "new_status_display": "Merged",
            })
            
            # Broadcast update for primary ticket so active detail/list pages reload
            _broadcast_ticket_update(primary_ticket, "ticket_status_changed", {
                "changed_by": request.user.username,
                "new_status": primary_ticket.status,
                "new_status_display": primary_ticket.get_status_display(),
            })
            
            return redirect("ticket_detail", ticket_id=primary_ticket.id)
        except ValidationError as e:
            for message in getattr(e, "messages", [str(e)]):
                django_messages.error(request, message)
        except Exception as e:
            django_messages.error(request, f"Failed to merge ticket: {str(e)}")

    return redirect("ticket_detail", ticket_id=ticket_id)
# Trigger Daphne Reload
